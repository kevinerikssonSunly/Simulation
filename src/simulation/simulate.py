import pandas as pd
from openpyxl import load_workbook
from typing import List
from models.resource import Wind, PV
from optimisation.search import simulate_dispatch_per_year
from utils.profiles import get_profiles

def run_all_scenarios(
    profiles: pd.DataFrame,
    wind_profile: pd.Series,
    solar_profile: pd.Series,
    baseload,
    output_path: str = "/Users/kevineriksson/PycharmProjects/Simulation/data/Simulator output.xlsx"
) -> None:
    """
    Run dispatch simulation for specified baseload and write the results to an Excel sheet.

    Args:
        wind_profile: Hourly wind profile (normalized or actual).
        solar_profile: Hourly solar profile (normalized or actual).
    """
    print(f"\nRunning simulation for {baseload} MW baseload...")

    wind_prod, solar_prod = get_profiles(44, 183, profiles)

    results, hourly_df = simulate_dispatch_per_year(
        profile_file=profiles,
        wind_prod=wind_prod,
        solar_prod=solar_prod,
        baseload=baseload,
        wind_cap=44,
        solar_cap=183,
        battery_1h_mw=0,
        battery_2h_mw=0,
        battery_4h_mw=150,
        battery_6h_mw=0,
        battery_8h_mw=0,
        hydro_mw=0,
        bess_rte=0.86,
        hydro_rte=0.9
    )

    df = pd.DataFrame(results)
    append_to_excel(df, output_path)

def append_to_excel(
    df: pd.DataFrame,
    excel_path: str,
    start_row: int = 5,
    start_col: int = 2
) -> None:
    """
    Append a DataFrame to the first empty row of an Excel sheet starting from a specified position.

    Args:
        df: DataFrame to append.
        excel_path: Path to the Excel workbook.
        start_row: Starting row to check for empty space (1-based index).
        start_col: Starting column to insert data (1-based index).
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    row = start_row
    while ws.cell(row=row, column=start_col).value is not None:
        row += 1

    for r_idx, record in enumerate(df.itertuples(index=False), start=row):
        for c_idx, value in enumerate(record, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(excel_path)